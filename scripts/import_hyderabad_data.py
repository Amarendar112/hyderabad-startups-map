#!/usr/bin/env python3
"""Load the real Hyderabad startup dataset from the downloaded CSV/XLSX files and write a TypeScript source file."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = Path(r"C:\Users\jaalt\Downloads\hyderabad_startups_list.csv")
XLSX_PATH = Path(r"C:\Users\jaalt\Downloads\hyderabad_startups_with_logos.xlsx")
OUTPUT_PATH = ROOT / "src" / "data" / "startups.ts"

AREA_COORDS = {
    "HITEC City": {"lat": 17.4504, "lng": 78.3808},
    "Knowledge City / T-Hub": {"lat": 17.4398, "lng": 78.3812},
    "Gachibowli": {"lat": 17.4401, "lng": 78.3489},
    "Madhapur": {"lat": 17.4483, "lng": 78.3915},
    "Financial District": {"lat": 17.4146, "lng": 78.3435},
    "Kondapur": {"lat": 17.4649, "lng": 78.3657},
    "Jubilee Hills": {"lat": 17.4319, "lng": 78.4073},
    "Banjara Hills": {"lat": 17.4156, "lng": 78.4347},
    "Begumpet": {"lat": 17.4447, "lng": 78.4664},
    "Uppal & East": {"lat": 17.3984, "lng": 78.5583},
    "Genome Valley / Uppal": {"lat": 17.3984, "lng": 78.5583},
    "Shamshabad & Outer": {"lat": 17.2369, "lng": 78.3656},
}

AREA_INFO = [
    {"name": "HITEC City", "lat": 17.4504, "lng": 78.3808, "zoom": 14, "description": "The iconic technology heart of Hyderabad, home to Cyber Towers, IT parks, and high-growth SaaS unicorns.", "icon": "Building2"},
    {"name": "Knowledge City / T-Hub", "lat": 17.4398, "lng": 78.3812, "zoom": 15, "description": "World's largest innovation campus (T-Hub 2.0), Image Tower, and cutting-edge deep tech hubs.", "icon": "Sparkles"},
    {"name": "Gachibowli", "lat": 17.4401, "lng": 78.3489, "zoom": 14, "description": "Major software district & university ecosystem (IIT-H, ISB, HCU) driving AI and Enterprise tech.", "icon": "Cpu"},
    {"name": "Madhapur", "lat": 17.4483, "lng": 78.3915, "zoom": 14, "description": "Vibrant startup hub with coworking spaces, accelerators, design agencies, and product teams.", "icon": "Zap"},
    {"name": "Financial District", "lat": 17.4146, "lng": 78.3435, "zoom": 14, "description": "Nanakramguda financial center housing global tech giants, fintech startups, and venture funds.", "icon": "Landmark"},
    {"name": "Kondapur", "lat": 17.4649, "lng": 78.3657, "zoom": 14, "description": "Rapidly growing cluster connecting HITEC City and Gachibowli with residential & tech hubs.", "icon": "Compass"},
    {"name": "Jubilee Hills", "lat": 17.4319, "lng": 78.4073, "zoom": 14, "description": "Upscale business district housing prominent angel investors, family offices, and founder headquarters.", "icon": "Award"},
    {"name": "Banjara Hills", "lat": 17.4156, "lng": 78.4347, "zoom": 14, "description": "Prime central location for healthcare innovators, luxury consumer brands, and venture studios.", "icon": "Briefcase"},
    {"name": "Begumpet", "lat": 17.4447, "lng": 78.4664, "zoom": 13, "description": "Central Hyderabad business hub featuring legacy IT companies and fintech enterprise headquarters.", "icon": "Globe"},
    {"name": "Uppal & East", "lat": 17.3984, "lng": 78.5583, "zoom": 13, "description": "East Hyderabad technology zone with hardware labs, bio-incubators, and manufacturing centers.", "icon": "Layers"},
]

INDUSTRY_MAP = {
    "AI & Machine Learning": "AI & Machine Learning",
    "HealthTech & BioTech": "HealthTech & BioTech",
    "HealthTech": "HealthTech",
    "E-Commerce & Consumer": "E-Commerce & Consumer",
    "SaaS & Enterprise": "SaaS & Enterprise",
    "SaaS": "SaaS & Enterprise",
    "FinTech & InsurTech": "FinTech & InsurTech",
    "FinTech": "FinTech",
    "EdTech": "EdTech",
    "DeepTech & Aerospace": "DeepTech & Aerospace",
    "CleanTech & EV": "CleanTech & EV",
    "CleanTech & Energy": "CleanTech & Energy",
    "AgriTech & FoodTech": "AgriTech & FoodTech",
    "FoodTech": "FoodTech",
    "Logistics & Mobility": "Logistics & Mobility",
    "Logistics & Supply Chain": "Logistics & Supply Chain",
    "Space Tech": "Space Tech",
    "DeepTech & AI": "DeepTech & AI",
}

STAGE_MAP = {
    "Idea / Stealth": "Idea / Stealth",
    "MVP / Early Stage": "MVP / Early Stage",
    "Growth / Scaling": "Growth / Scaling",
    "Established / Unicorn": "Established / Unicorn",
    "Established / Public": "Established / Public",
    "Established / Acquired": "Established / Acquired",
    "Established / Profitable": "Established / Profitable",
    "Early Stage": "Early Stage",
}

FUNDING_MAP = {
    "Bootstrapped": "Bootstrapped",
    "Pre-Seed": "Pre-Seed",
    "Seed": "Seed",
    "Series A": "Series A",
    "Series B": "Series B",
    "Series C": "Series C",
    "Series C+": "Series C+",
    "Series D": "Series D",
    "Series E": "Series E",
    "Series F": "Series F",
    "IPO": "IPO",
    "Grants & Government": "Grants & Government",
    "Acquired": "Acquired",
    "Private Equity": "Private Equity",
    "Pre-Series A": "Pre-Series A",
}


def clean_string(value):
    if value is None:
        return ""
    return str(value).strip()


def slugify(value):
    return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-") or "startup"


def normalize_area(value):
    name = clean_string(value)
    if name in AREA_COORDS:
        return name
    for key in AREA_COORDS:
        if key.lower() == name.lower():
            return key
    return "HITEC City"


def parse_decimal(value):
    if value is None or value == "":
        return 0
    s = str(value).replace(",", "").replace("$", "").strip()
    if not s:
        return 0
    try:
        return float(s)
    except ValueError:
        return 0


def load_xlsx_map():
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [clean_string(cell) for cell in rows[0]]
    data = {}
    for row in rows[1:]:
        if not row or not any(clean_string(cell) for cell in row):
            continue
        record = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
        startup_id = clean_string(record.get("ID") or record.get("id") or record.get("Company ID"))
        if startup_id:
            data[startup_id] = record
    return data


def load_csv_rows():
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def build_startup_objects():
    xlsx_map = load_xlsx_map()
    csv_rows = load_csv_rows()
    startups = []

    for row in csv_rows:
        startup_id = clean_string(row.get("id") or row.get("ID"))
        if not startup_id:
            continue
        xlsx_row = xlsx_map.get(startup_id, {})

        name = clean_string(xlsx_row.get("Company Name") or row.get("name"))
        website = clean_string(row.get("website") or xlsx_row.get("Website") or "")
        if not website and xlsx_row.get("Website"):
            website = clean_string(xlsx_row.get("Website"))
        if website and not website.startswith("http"):
            website = "https://" + website

        domain = clean_string(xlsx_row.get("Domain") or (website.split("//")[-1].split('/')[0] if website else ""))
        google_logo = clean_string(xlsx_row.get("Logo URL (Google - primary, working)") or "")
        dev_logo = clean_string(xlsx_row.get("Logo URL (logo.dev - pending new key)") or "")
        fallback_svg = clean_string(xlsx_row.get("Fallback Avatar (SVG data URI)") or "")
        raw_logo = google_logo or dev_logo or ""

        industry_raw = clean_string(xlsx_row.get("Industry") or row.get("industry") or "SaaS & Enterprise")
        industry = INDUSTRY_MAP.get(industry_raw, "SaaS & Enterprise")

        area = normalize_area(xlsx_row.get("Area") or row.get("area") or "HITEC City")
        coords = AREA_COORDS.get(area, {"lat": 17.4435, "lng": 78.3772})

        startup = {
            "id": startup_id,
            "name": name,
            "slug": slugify(name or startup_id),
            "tagline": f"Hyderabad startup in {industry}.",
            "description": f"{name or 'This Hyderabad startup'} operates in the {industry} ecosystem and is part of the city’s startup landscape.",
            "logoUrl": raw_logo,
            "website": website,
            "industry": industry,
            "subCategory": "Hyderabad Startup",
            "stage": "MVP / Early Stage",
            "fundingStage": "Bootstrapped",
            "totalFunding": "Not disclosed",
            "totalFundingAmountUsd": 0,
            "foundingYear": 0,
            "teamSize": "Not disclosed",
            "location": {
                "area": area,
                "address": f"{area}, Hyderabad",
                "lat": coords["lat"],
                "lng": coords["lng"],
                "building": area,
            },
            "founders": [],
            "hiring": bool(clean_string(row.get("careers_url") or website or "")),
            "hiringRoles": [],
            "tags": [industry, area, "Hyderabad"],
            "featured": False,
            "verified": True,
            "incubationHub": "",
            "createdAt": "2026-08-30",
            "updatedAt": "2026-08-30",
            "jobOpenings": [],
            "svgAvatar": fallback_svg,
        }
        startups.append(startup)

    return startups


def render_ts(startups):
    lines = [
        "import { Startup, Incubator, Investor, HyderabadAreaInfo } from '@/types/startup';",
        "",
        "export const HYDERABAD_AREAS: HyderabadAreaInfo[] = [",
    ]

    for item in AREA_INFO:
        lines.append("  {")
        lines.append(f"    name: '{item['name']}',")
        lines.append(f"    lat: {item['lat']},")
        lines.append(f"    lng: {item['lng']},")
        lines.append(f"    zoom: {item['zoom']},")
        lines.append(f"    description: '{item['description'].replace("'", "\\'")}',")
        lines.append(f"    icon: '{item['icon']}',")
        lines.append("  },")

    lines.append("];")
    lines.append("")
    lines.append("export const INITIAL_STARTUPS: Startup[] = [")

    for i, startup in enumerate(startups):
        lines.append("  {")
        lines.append(f"    id: '{startup['id']}',")
        lines.append(f"    name: '{startup['name'].replace("'", "\\'")}',")
        lines.append(f"    slug: '{startup['slug']}',")
        lines.append(f"    tagline: '{startup['tagline'].replace("'", "\\'")}',")
        lines.append(f"    description: '{startup['description'].replace("'", "\\'")}',")
        lines.append(f"    logoUrl: '{startup['logoUrl'].replace("'", "\\'")}',")
        lines.append(f"    website: '{startup['website'].replace("'", "\\'")}',")
        lines.append(f"    industry: '{startup['industry']}',")
        lines.append(f"    subCategory: '{startup['subCategory'].replace("'", "\\'")}',")
        lines.append(f"    stage: '{startup['stage']}',")
        lines.append(f"    fundingStage: '{startup['fundingStage']}',")
        lines.append(f"    totalFunding: '{startup['totalFunding'].replace("'", "\\'")}',")
        lines.append(f"    totalFundingAmountUsd: {startup['totalFundingAmountUsd']},")
        lines.append(f"    foundingYear: {startup['foundingYear']},")
        lines.append(f"    teamSize: '{startup['teamSize'].replace("'", "\\'")}',")
        lines.append("    location: {")
        lines.append(f"      area: '{startup['location']['area']}',")
        lines.append(f"      address: '{startup['location']['address'].replace("'", "\\'")}',")
        lines.append(f"      lat: {startup['location']['lat']},")
        lines.append(f"      lng: {startup['location']['lng']},")
        lines.append(f"      building: '{startup['location']['building'].replace("'", "\\'")}',")
        lines.append("    },")
        lines.append("    founders: [],")
        lines.append(f"    hiring: {str(startup['hiring']).lower()},")
        lines.append("    hiringRoles: [],")
        lines.append("    tags: [")
        for tag in startup["tags"]:
            lines.append(f"      '{tag.replace("'", "\\'")}',")
        lines.append("    ],")
        lines.append(f"    featured: {str(startup['featured']).lower()},")
        lines.append(f"    verified: {str(startup['verified']).lower()},")
        lines.append(f"    createdAt: '{startup['createdAt']}',")
        lines.append(f"    updatedAt: '{startup['updatedAt']}',")
        lines.append("    jobOpenings: [],")
        if startup.get("svgAvatar"):
            lines.append(f"    svgAvatar: '{startup['svgAvatar'].replace("'", "\\'")}',")
        lines.append("  },")

    lines.append("];")
    lines.append("")
    lines.append("export const INITIAL_INCUBATORS: Incubator[] = [];")
    lines.append("")
    lines.append("export const INITIAL_INVESTORS: Investor[] = [];")
    return "\n".join(lines) + "\n"


def main():
    startups = build_startup_objects()
    output = render_ts(startups)
    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"Wrote {len(startups)} startups to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
